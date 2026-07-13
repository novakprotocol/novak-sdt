#!/usr/bin/env python3
"""Build a reproducible English lexical-form inventory from a full Kaikki JSONL dump.

Acceptance is source-record based: a row must be an English, single-token lexical item
with either an independent gloss or an explicit form-of/alternative-of relationship.
Proper names, phrases, affixes, symbols, punctuation, documented misspellings,
eye-dialect spellings, reconstructed forms, and taxonomic names are excluded.

DWYL and Webster datasets are used only as reconnaissance flags. They never determine
acceptance or ranking.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import gzip
import hashlib
import json
import os
import re
import shutil
import sqlite3
import sys
import unicodedata
import urllib.request
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Iterator

try:
    import orjson  # type: ignore
except Exception:  # pragma: no cover
    orjson = None

try:
    import xlsxwriter  # type: ignore
except Exception as exc:  # pragma: no cover
    raise SystemExit("xlsxwriter is required") from exc

try:
    import openpyxl  # type: ignore
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required") from exc

VERSION = "2.0.0"
TARGET = 1_000_000
SOURCE_NAME = "Kaikki English dictionary (Wiktionary/Wiktextract extraction)"
DWYL_URL = "https://raw.githubusercontent.com/dwyl/english-words/master/words_alpha.txt"
WEBSTER_REPO_API = "https://api.github.com/repos/ssvivian/WebstersDictionary"

APOSTROPHES = {"’", "‘", "ʼ", "ʻ", "ʹ", "`", "´", "＇"}
DASHES = {"‐", "‑", "‒", "–", "—", "―", "−", "﹘", "﹣", "－"}

EXCLUDED_POS = {
    "name", "proper noun", "proper_noun", "phrase", "proverb", "prefix", "suffix",
    "infix", "circumfix", "interfix", "affix", "combining form", "combining_form",
    "character", "letter", "symbol", "punct", "punctuation", "root", "romanization",
}

HARD_REJECT_TAG_FRAGMENTS = {
    "misspelling", "misconstruction", "eye dialect", "eye-dialect",
    "deliberate misspelling", "common misspelling", "reconstruction", "reconstructed",
    "taxonomic", "proper noun", "proper-noun", "proper name", "proper-name",
}

STATUS_TAGS = {
    "archaic", "obsolete", "historical", "rare", "dated", "dialectal", "regional",
    "informal", "colloquial", "slang", "nonstandard", "technical", "chiefly",
}

COMMON_POS = {
    "noun", "verb", "adj", "adjective", "adv", "adverb", "pron", "pronoun",
    "det", "determiner", "prep", "preposition", "conj", "conjunction", "interj",
    "interjection", "num", "numeral", "particle", "article", "contraction",
}

ABBREV_POS = {"abbrev", "abbreviation", "acronym", "initialism"}

CSV_FIELDS = [
    "entry", "canonical_entry", "entry_class", "part_of_speech", "definition",
    "form_of", "usage_status", "evidence_type", "source_record", "source_line",
    "source_evidence_count", "quality_score", "recon_dwyl", "recon_webster",
]


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def parse_json_line(raw: bytes) -> dict[str, Any]:
    if orjson is not None:
        return orjson.loads(raw)
    return json.loads(raw)


def normalize_display(value: str) -> str:
    value = unicodedata.normalize("NFKC", value.strip())
    value = "".join("'" if c in APOSTROPHES else "-" if c in DASHES else c for c in value)
    return value


def canonicalize(value: str) -> str:
    return normalize_display(value).casefold()


def token_shape_reason(word: str) -> str | None:
    if not word:
        return "empty"
    if len(word) > 100:
        return "over_100_codepoints"
    if any(ch.isspace() for ch in word):
        return "contains_whitespace"
    if word[0] in "-'" or word[-1] in "-'":
        return "edge_joiner"
    if "--" in word or "''" in word or "-'" in word or "'-" in word:
        return "repeated_joiner"
    letters = 0
    for ch in word:
        cat = unicodedata.category(ch)
        if cat.startswith("L"):
            letters += 1
        elif cat.startswith("M"):
            continue
        elif ch in "-'":
            continue
        else:
            return "nonlexical_character"
    if letters == 0:
        return "no_letters"
    return None


def norm_tag(tag: Any) -> str:
    return re.sub(r"[_-]+", " ", str(tag).strip().casefold())


def flatten_tags(obj: dict[str, Any], sense: dict[str, Any] | None = None) -> set[str]:
    values: list[Any] = []
    values.extend(obj.get("tags") or [])
    if sense:
        values.extend(sense.get("tags") or [])
    return {norm_tag(x) for x in values if str(x).strip()}


def hard_reject_tag(tags: Iterable[str]) -> str | None:
    for tag in tags:
        for fragment in HARD_REJECT_TAG_FRAGMENTS:
            if fragment in tag:
                return fragment.replace(" ", "_")
    return None


def first_text(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        for item in value:
            text = first_text(item)
            if text:
                return text
    return ""


def relation_words(value: Any) -> list[str]:
    out: list[str] = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                candidate = item.get("word") or item.get("name") or item.get("term")
                if isinstance(candidate, str) and candidate.strip():
                    out.append(normalize_display(candidate))
            elif isinstance(item, str) and item.strip():
                out.append(normalize_display(item))
    elif isinstance(value, dict):
        candidate = value.get("word") or value.get("name") or value.get("term")
        if isinstance(candidate, str) and candidate.strip():
            out.append(normalize_display(candidate))
    return list(dict.fromkeys(out))


def select_evidence(obj: dict[str, Any]) -> tuple[dict[str, Any] | None, set[str]]:
    senses = obj.get("senses") or []
    if not isinstance(senses, list):
        senses = []
    options: list[dict[str, Any]] = []
    all_reasons: set[str] = set()
    for sense in senses:
        if not isinstance(sense, dict):
            continue
        tags = flatten_tags(obj, sense)
        tag_reason = hard_reject_tag(tags)
        if tag_reason:
            all_reasons.add(f"tag_{tag_reason}")
            continue
        gloss = first_text(sense.get("glosses")) or first_text(sense.get("raw_glosses"))
        if gloss.casefold() in {"no definition", "no gloss", "unknown", "?"}:
            gloss = ""
        form_of = relation_words(sense.get("form_of"))
        alt_of = relation_words(sense.get("alt_of"))
        if not gloss and not form_of and not alt_of:
            all_reasons.add("no_gloss_or_relation")
            continue
        evidence_type = "gloss"
        relation = ""
        if form_of:
            evidence_type = "form_of"
            relation = "; ".join(form_of[:12])
        elif alt_of:
            evidence_type = "alternative_of"
            relation = "; ".join(alt_of[:12])
        score = 1000 if gloss and not (form_of or alt_of) else 850 if alt_of else 760
        if gloss:
            score += min(len(gloss), 300) // 10
        if tags & {"archaic", "obsolete", "dated", "historical"}:
            score -= 45
        if tags & {"rare", "nonstandard"}:
            score -= 20
        options.append({
            "definition": gloss[:2000],
            "form_of": relation,
            "evidence_type": evidence_type,
            "tags": tags,
            "score": score,
        })
    if not options:
        return None, all_reasons or {"no_eligible_sense"}
    options.sort(key=lambda x: (x["score"], len(x["definition"])), reverse=True)
    return options[0], all_reasons


def usage_status(tags: set[str]) -> str:
    selected = sorted({t for t in tags if any(key in t for key in STATUS_TAGS)})
    return "; ".join(selected[:20]) if selected else "current_or_unmarked"


def classify(pos: str, evidence_type: str) -> str:
    if evidence_type == "form_of":
        return "documented_inflection_or_form"
    if evidence_type == "alternative_of":
        return "documented_variant"
    if pos in ABBREV_POS:
        return "defined_abbreviation"
    if pos == "contraction":
        return "contraction"
    return "independently_glossed_lexeme"


def merge_semicolon(old: str, new: str, limit: int = 40) -> str:
    values = [x.strip() for x in (old + ";" + new).split(";") if x.strip()]
    return "; ".join(list(dict.fromkeys(values))[:limit])


def init_db(path: Path) -> sqlite3.Connection:
    if path.exists():
        path.unlink()
    db = sqlite3.connect(path)
    db.executescript(
        """
        PRAGMA journal_mode=OFF;
        PRAGMA synchronous=OFF;
        PRAGMA temp_store=FILE;
        PRAGMA locking_mode=EXCLUSIVE;
        PRAGMA cache_size=-262144;
        CREATE TABLE accepted (
          canonical TEXT PRIMARY KEY,
          entry TEXT NOT NULL,
          entry_class TEXT NOT NULL,
          pos TEXT NOT NULL,
          definition TEXT NOT NULL,
          form_of TEXT NOT NULL,
          usage_status TEXT NOT NULL,
          evidence_type TEXT NOT NULL,
          source_record TEXT NOT NULL,
          source_line INTEGER NOT NULL,
          evidence_count INTEGER NOT NULL,
          score INTEGER NOT NULL,
          recon_dwyl INTEGER NOT NULL DEFAULT 0,
          recon_webster INTEGER NOT NULL DEFAULT 0,
          in_million INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE rejected (
          canonical TEXT PRIMARY KEY,
          entry TEXT NOT NULL,
          reason TEXT NOT NULL,
          pos TEXT NOT NULL,
          source_record TEXT NOT NULL,
          source_line INTEGER NOT NULL
        );
        CREATE TABLE reason_counts (reason TEXT PRIMARY KEY, count INTEGER NOT NULL);
        """
    )
    return db


def increment_reason(db: sqlite3.Connection, reason: str) -> None:
    db.execute(
        "INSERT INTO reason_counts(reason,count) VALUES(?,1) "
        "ON CONFLICT(reason) DO UPDATE SET count=count+1", (reason,)
    )


def reject(db: sqlite3.Connection, canonical: str, entry: str, reason: str, pos: str,
           record: str, line_no: int) -> None:
    increment_reason(db, reason)
    if canonical:
        db.execute(
            "INSERT OR IGNORE INTO rejected(canonical,entry,reason,pos,source_record,source_line) "
            "VALUES(?,?,?,?,?,?)", (canonical, entry, reason, pos, record, line_no)
        )


def upsert_accepted(db: sqlite3.Connection, row: dict[str, Any]) -> None:
    current = db.execute(
        "SELECT entry_class,pos,definition,form_of,usage_status,evidence_type,source_record,"
        "source_line,evidence_count,score,entry FROM accepted WHERE canonical=?",
        (row["canonical"],),
    ).fetchone()
    if current is None:
        db.execute(
            "INSERT INTO accepted(canonical,entry,entry_class,pos,definition,form_of,usage_status,"
            "evidence_type,source_record,source_line,evidence_count,score) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,1,?)",
            (row["canonical"], row["entry"], row["entry_class"], row["pos"],
             row["definition"], row["form_of"], row["usage_status"], row["evidence_type"],
             row["source_record"], row["source_line"], row["score"]),
        )
    else:
        (old_class, old_pos, old_def, old_form, old_status, old_ev, old_record,
         old_line, old_count, old_score, old_entry) = current
        merged_pos = merge_semicolon(old_pos, row["pos"])
        merged_status = merge_semicolon(old_status, row["usage_status"])
        if row["score"] > old_score:
            values = (row["entry"], row["entry_class"], merged_pos, row["definition"],
                      row["form_of"], merged_status, row["evidence_type"], row["source_record"],
                      row["source_line"], old_count + 1, row["score"], row["canonical"])
        else:
            values = (old_entry, old_class, merged_pos, old_def, old_form, merged_status,
                      old_ev, old_record, old_line, old_count + 1, old_score, row["canonical"])
        db.execute(
            "UPDATE accepted SET entry=?,entry_class=?,pos=?,definition=?,form_of=?,usage_status=?,"
            "evidence_type=?,source_record=?,source_line=?,evidence_count=?,score=? WHERE canonical=?",
            values,
        )
    db.execute("DELETE FROM rejected WHERE canonical=?", (row["canonical"],))


def process_jsonl(source: Path, db: sqlite3.Connection, progress_path: Path) -> dict[str, Any]:
    counters: Counter[str] = Counter()
    last_commit = 0
    with source.open("rb") as f:
        for line_no, raw in enumerate(f, 1):
            if not raw.strip():
                continue
            counters["source_lines"] += 1
            record_hash = hashlib.sha256(raw).hexdigest()[:24]
            try:
                obj = parse_json_line(raw)
            except Exception:
                counters["json_errors"] += 1
                increment_reason(db, "invalid_json")
                continue
            if obj.get("lang_code") != "en":
                counters["non_english_records"] += 1
                continue
            raw_word = obj.get("word")
            if not isinstance(raw_word, str):
                counters["missing_word"] += 1
                increment_reason(db, "missing_word")
                continue
            entry = normalize_display(raw_word)
            canonical = canonicalize(entry)
            pos = norm_tag(obj.get("pos") or obj.get("pos_title") or "unknown")
            shape_reason = token_shape_reason(entry)
            if shape_reason:
                reject(db, canonical, entry, shape_reason, pos, record_hash, line_no)
                counters["rejected_records"] += 1
                continue
            if pos in EXCLUDED_POS or any(x in pos for x in ("proper noun", "proper name")):
                reject(db, canonical, entry, "excluded_pos_" + pos.replace(" ", "_"), pos,
                       record_hash, line_no)
                counters["rejected_records"] += 1
                continue
            obj_tags = flatten_tags(obj)
            tag_reason = hard_reject_tag(obj_tags)
            if tag_reason:
                reject(db, canonical, entry, "tag_" + tag_reason.replace(" ", "_"), pos,
                       record_hash, line_no)
                counters["rejected_records"] += 1
                continue
            evidence, _ = select_evidence(obj)
            if evidence is None:
                reject(db, canonical, entry, "no_eligible_gloss_or_relation", pos,
                       record_hash, line_no)
                counters["rejected_records"] += 1
                continue
            tags = obj_tags | set(evidence["tags"])
            score = int(evidence["score"])
            if pos in COMMON_POS:
                score += 35
            if pos in ABBREV_POS:
                score -= 30
            row = {
                "canonical": canonical,
                "entry": entry,
                "entry_class": classify(pos, evidence["evidence_type"]),
                "pos": pos,
                "definition": evidence["definition"],
                "form_of": evidence["form_of"],
                "usage_status": usage_status(tags),
                "evidence_type": evidence["evidence_type"],
                "source_record": f"line:{line_no};sha256:{record_hash}",
                "source_line": line_no,
                "score": score,
            }
            upsert_accepted(db, row)
            counters["accepted_records"] += 1
            if line_no - last_commit >= 20_000:
                db.commit()
                last_commit = line_no
                progress_path.write_text(json.dumps({**counters, "line": line_no, "at": utc_now()}, indent=2))
    db.commit()
    counters["accepted_unique"] = db.execute("SELECT COUNT(*) FROM accepted").fetchone()[0]
    counters["rejected_unique"] = db.execute("SELECT COUNT(*) FROM rejected").fetchone()[0]
    return dict(counters)


def url_bytes(url: str, timeout: int = 120) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "EnglishLexiconProof/2.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def download_recon_dwyl(cache: Path) -> set[str]:
    path = cache / "dwyl_words_alpha.txt"
    if not path.exists():
        path.write_bytes(url_bytes(DWYL_URL))
    result: set[str] = set()
    with path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            word = line.strip()
            if word:
                result.add(canonicalize(word))
    return result


def find_webster_download() -> tuple[str, str]:
    repo = json.loads(url_bytes(WEBSTER_REPO_API))
    branch = repo.get("default_branch") or "master"
    tree_url = f"{WEBSTER_REPO_API}/git/trees/{branch}?recursive=1"
    tree = json.loads(url_bytes(tree_url)).get("tree", [])
    candidates = []
    for item in tree:
        path = item.get("path", "")
        size = int(item.get("size") or 0)
        if item.get("type") == "blob" and path.lower().endswith((".json", ".csv", ".txt")):
            name = path.lower()
            priority = 3 if "dictionary" in name and name.endswith(".json") else 2 if name.endswith(".json") else 1
            candidates.append((priority, size, path))
    if not candidates:
        raise RuntimeError("No Webster data file discovered")
    _, _, path = max(candidates)
    return f"https://raw.githubusercontent.com/ssvivian/WebstersDictionary/{branch}/{path}", path


def collect_strings_from_json(value: Any, out: set[str], depth: int = 0) -> None:
    if depth > 4:
        return
    if isinstance(value, dict):
        for k, v in value.items():
            if isinstance(k, str) and k.strip() and len(k) <= 100:
                if token_shape_reason(normalize_display(k)) is None:
                    out.add(canonicalize(k))
            if isinstance(v, dict):
                for key in ("word", "entry", "headword", "term"):
                    candidate = v.get(key)
                    if isinstance(candidate, str) and token_shape_reason(normalize_display(candidate)) is None:
                        out.add(canonicalize(candidate))
                collect_strings_from_json(v, out, depth + 1)
            elif isinstance(v, list):
                collect_strings_from_json(v, out, depth + 1)
    elif isinstance(value, list):
        for item in value:
            collect_strings_from_json(item, out, depth + 1)


def download_recon_webster(cache: Path) -> tuple[set[str], str]:
    url, relpath = find_webster_download()
    suffix = Path(relpath).suffix.lower()
    path = cache / ("webster_source" + suffix)
    if not path.exists():
        path.write_bytes(url_bytes(url, timeout=300))
    result: set[str] = set()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        collect_strings_from_json(data, result)
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():
                    w = normalize_display(row[0])
                    if token_shape_reason(w) is None:
                        result.add(canonicalize(w))
    else:
        with path.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                w = normalize_display(line.strip().split("\t", 1)[0])
                if w and token_shape_reason(w) is None:
                    result.add(canonicalize(w))
    return result, url


def apply_recon(db: sqlite3.Connection, dwyl: set[str], webster: set[str]) -> dict[str, int]:
    dwyl_hits = 0
    webster_hits = 0
    cur = db.execute("SELECT canonical FROM accepted")
    updates = []
    for (canonical,) in cur:
        d = int(canonical in dwyl)
        w = int(canonical in webster)
        dwyl_hits += d
        webster_hits += w
        if d or w:
            updates.append((d, w, canonical))
        if len(updates) >= 50_000:
            db.executemany("UPDATE accepted SET recon_dwyl=?,recon_webster=? WHERE canonical=?", updates)
            updates.clear()
    if updates:
        db.executemany("UPDATE accepted SET recon_dwyl=?,recon_webster=? WHERE canonical=?", updates)
    db.commit()
    return {"dwyl_hits": dwyl_hits, "webster_hits": webster_hits,
            "dwyl_inventory": len(dwyl), "webster_inventory": len(webster)}


def mark_million(db: sqlite3.Connection, target: int) -> int:
    db.execute("UPDATE accepted SET in_million=0")
    count = db.execute("SELECT COUNT(*) FROM accepted").fetchone()[0]
    chosen = min(count, target)
    db.execute(
        "UPDATE accepted SET in_million=1 WHERE canonical IN ("
        "SELECT canonical FROM accepted ORDER BY score DESC, "
        "CASE evidence_type WHEN 'gloss' THEN 0 WHEN 'alternative_of' THEN 1 ELSE 2 END, "
        "canonical ASC LIMIT ?)", (chosen,)
    )
    db.commit()
    return chosen


def initial_bucket(entry: str) -> str:
    if not entry:
        return "OTHER"
    first = unicodedata.normalize("NFKD", entry[0]).encode("ascii", "ignore").decode().upper()
    return first[0] if first and "A" <= first[0] <= "Z" else "OTHER"


def row_from_sql(row: tuple[Any, ...]) -> dict[str, Any]:
    return dict(zip(CSV_FIELDS, row))


def accepted_query(where: str = "") -> str:
    return (
        "SELECT entry,canonical,entry_class,pos,definition,form_of,usage_status,evidence_type,"
        "source_record,source_line,evidence_count,score,recon_dwyl,recon_webster FROM accepted "
        + where + " ORDER BY canonical"
    )


def export_partitions(db: sqlite3.Connection, out_root: Path, where: str) -> tuple[int, dict[str, int]]:
    out_root.mkdir(parents=True, exist_ok=True)
    handles: dict[str, Any] = {}
    writers: dict[str, csv.DictWriter] = {}
    counts: Counter[str] = Counter()
    total = 0
    try:
        for bucket in list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["OTHER"]:
            handle = gzip.open(out_root / f"{bucket}.csv.gz", "wt", encoding="utf-8", newline="", compresslevel=6)
            writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS, extrasaction="ignore")
            writer.writeheader()
            handles[bucket] = handle
            writers[bucket] = writer
        for sql_row in db.execute(accepted_query(where)):
            row = row_from_sql(sql_row)
            bucket = initial_bucket(str(row["entry"]))
            writers[bucket].writerow(row)
            counts[bucket] += 1
            total += 1
    finally:
        for handle in handles.values():
            handle.close()
    return total, dict(counts)


def export_rejected(db: sqlite3.Connection, path: Path) -> int:
    fields = ["entry", "canonical_entry", "rejection_reason", "part_of_speech", "source_record", "source_line"]
    count = 0
    with gzip.open(path, "wt", encoding="utf-8", newline="", compresslevel=6) as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for canonical, entry, reason, pos, record, line_no in db.execute(
            "SELECT canonical,entry,reason,pos,source_record,source_line FROM rejected ORDER BY canonical"
        ):
            writer.writerow([entry, canonical, reason, pos, record, line_no])
            count += 1
    return count


def sql_counts(db: sqlite3.Connection, million_only: bool) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    condition = "WHERE in_million=1" if million_only else ""
    letters: Counter[str] = Counter()
    pos: Counter[str] = Counter()
    classes: Counter[str] = Counter()
    for entry, p, cls in db.execute(f"SELECT entry,pos,entry_class FROM accepted {condition}"):
        letters[initial_bucket(entry)] += 1
        for item in p.split(";"):
            if item.strip():
                pos[item.strip()] += 1
        classes[cls] += 1
    return dict(letters), dict(pos), dict(classes)


def write_workbook(db: sqlite3.Connection, path: Path, metadata: dict[str, Any], target_rows: int) -> None:
    workbook = xlsxwriter.Workbook(path, {"constant_memory": True, "strings_to_urls": False})
    workbook.set_properties({
        "title": "Verified English Lexical Forms — Kaikki/Wiktionary",
        "subject": "Source-attested English lexical forms",
        "author": "Reproducible lexicon pipeline",
        "comments": "Generated from a pinned Kaikki/Wiktionary source snapshot.",
    })
    fmt_title = workbook.add_format({"bold": True, "font_size": 18})
    fmt_h1 = workbook.add_format({"bold": True, "font_size": 12, "bg_color": "#D9EAF7", "border": 1})
    fmt_head = workbook.add_format({"bold": True, "bg_color": "#1F4E78", "font_color": "#FFFFFF", "border": 1, "text_wrap": True})
    fmt_wrap = workbook.add_format({"text_wrap": True, "valign": "top"})
    fmt_int = workbook.add_format({"num_format": "0"})

    readme = workbook.add_worksheet("README")
    readme.set_column("A:A", 26)
    readme.set_column("B:B", 110)
    readme.write("A1", "English Lexicon Verification Package", fmt_title)
    readme.write("A3", "Claim", fmt_h1)
    readme.write("B3", "Unique, source-attested English lexical forms with an independent gloss or explicit form relationship; complete against the documented source snapshot and rules.", fmt_wrap)
    readme.write("A4", "Not claimed", fmt_h1)
    readme.write("B4", "Not every possible English word, and not one million independent lemmas. Proper names, phrases, affixes, symbols, punctuation, misspellings, eye-dialect, reconstructions, and taxonomic names are excluded.", fmt_wrap)
    readme.write("A5", "Reconnaissance", fmt_h1)
    readme.write("B5", "DWYL and Webster flags are comparison fields only. They never validate or rank an entry.", fmt_wrap)
    readme.write("A7", "Verification", fmt_h1)
    readme.write("B7", "See proof/verification_report.json and proof/file_hashes.sha256 in the companion ZIP.", fmt_wrap)

    summary = workbook.add_worksheet("Summary")
    summary.set_column("A:A", 34)
    summary.set_column("B:B", 80)
    summary.write_row(0, 0, ["Metric", "Value"], fmt_head)
    for r, (key, value) in enumerate(metadata.items(), 1):
        summary.write(r, 0, str(key))
        summary.write(r, 1, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value)

    letters, pos_counts, class_counts = sql_counts(db, million_only=True)
    ws_letters = workbook.add_worksheet("Counts_By_Letter")
    ws_letters.write_row(0, 0, ["Initial", "Rows"], fmt_head)
    for i, key in enumerate(sorted(letters), 1):
        ws_letters.write(i, 0, key)
        ws_letters.write(i, 1, letters[key], fmt_int)
    ws_letters.set_column("A:A", 12)
    ws_letters.set_column("B:B", 16)

    ws_pos = workbook.add_worksheet("Counts_By_POS")
    ws_pos.write_row(0, 0, ["Part of speech", "Rows"], fmt_head)
    for i, (key, value) in enumerate(sorted(pos_counts.items(), key=lambda x: (-x[1], x[0])), 1):
        ws_pos.write(i, 0, key)
        ws_pos.write(i, 1, value, fmt_int)
    ws_pos.set_column("A:A", 32)
    ws_pos.set_column("B:B", 16)

    ws_class = workbook.add_worksheet("Counts_By_Class")
    ws_class.write_row(0, 0, ["Entry class", "Rows"], fmt_head)
    for i, (key, value) in enumerate(sorted(class_counts.items(), key=lambda x: (-x[1], x[0])), 1):
        ws_class.write(i, 0, key)
        ws_class.write(i, 1, value, fmt_int)
    ws_class.set_column("A:A", 38)
    ws_class.set_column("B:B", 16)

    data = workbook.add_worksheet("Million_Lexicon")
    headers = [
        "entry", "canonical_entry", "entry_class", "part_of_speech", "definition",
        "form_of", "usage_status", "evidence_type", "source_record", "source_evidence_count",
        "quality_score", "recon_dwyl", "recon_webster",
    ]
    data.write_row(0, 0, headers, fmt_head)
    widths = [28, 28, 34, 24, 90, 35, 30, 18, 34, 22, 16, 14, 16]
    for col, width in enumerate(widths):
        data.set_column(col, col, width)
    data.freeze_panes(1, 0)
    data.autofilter(0, 0, target_rows, len(headers) - 1)
    row_index = 1
    query = accepted_query("WHERE in_million=1")
    for sql_row in db.execute(query):
        row = row_from_sql(sql_row)
        values = [
            row["entry"], row["canonical_entry"], row["entry_class"], row["part_of_speech"],
            row["definition"], row["form_of"], row["usage_status"], row["evidence_type"],
            row["source_record"], row["source_evidence_count"], row["quality_score"],
            row["recon_dwyl"], row["recon_webster"],
        ]
        data.write_row(row_index, 0, values)
        row_index += 1
    if row_index - 1 != target_rows:
        raise RuntimeError(f"Workbook row mismatch: {row_index - 1} != {target_rows}")
    workbook.close()


def count_gzip_csv_rows(path: Path) -> int:
    with gzip.open(path, "rt", encoding="utf-8", newline="") as f:
        return max(sum(1 for _ in f) - 1, 0)


def verify_workbook(path: Path, expected: int) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        required = {"README", "Summary", "Counts_By_Letter", "Counts_By_POS", "Counts_By_Class", "Million_Lexicon"}
        missing = required - set(wb.sheetnames)
        if missing:
            raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
        ws = wb["Million_Lexicon"]
        rows = 0
        header = None
        for values in ws.iter_rows(values_only=True):
            if header is None:
                header = list(values)
            else:
                rows += 1
        if rows != expected:
            raise RuntimeError(f"Workbook data rows {rows} != {expected}")
        return {"sheet_names": wb.sheetnames, "million_rows": rows, "header": header}
    finally:
        wb.close()


def verify_unique(db: sqlite3.Connection, target: int) -> dict[str, Any]:
    total = db.execute("SELECT COUNT(*) FROM accepted WHERE in_million=1").fetchone()[0]
    distinct = db.execute("SELECT COUNT(DISTINCT canonical) FROM accepted WHERE in_million=1").fetchone()[0]
    missing = db.execute(
        "SELECT COUNT(*) FROM accepted WHERE in_million=1 AND "
        "((definition='' OR definition IS NULL) AND (form_of='' OR form_of IS NULL))"
    ).fetchone()[0]
    excluded_pos = db.execute(
        "SELECT COUNT(*) FROM accepted WHERE in_million=1 AND pos IN (" + ",".join("?" for _ in EXCLUDED_POS) + ")",
        tuple(EXCLUDED_POS),
    ).fetchone()[0]
    if total != target or distinct != target or missing or excluded_pos:
        raise RuntimeError({"total": total, "distinct": distinct, "missing_evidence": missing, "excluded_pos": excluded_pos})
    return {"total": total, "distinct_canonical": distinct, "missing_evidence": missing, "excluded_pos": excluded_pos}


def write_file_hashes(package: Path) -> dict[str, str]:
    proof = package / "proof"
    proof.mkdir(exist_ok=True)
    hash_path = proof / "file_hashes.sha256"
    hashes: dict[str, str] = {}
    for path in sorted(p for p in package.rglob("*") if p.is_file() and p != hash_path):
        rel = path.relative_to(package).as_posix()
        hashes[rel] = sha256_file(path)
    hash_path.write_text("".join(f"{digest}  {rel}\n" for rel, digest in hashes.items()), encoding="utf-8")
    return hashes


def make_zip(package: Path, out_zip: Path) -> None:
    if out_zip.exists():
        out_zip.unlink()
    with zipfile.ZipFile(out_zip, "w", allowZip64=True) as zf:
        for path in sorted(p for p in package.rglob("*") if p.is_file()):
            rel = path.relative_to(package.parent).as_posix()
            compression = zipfile.ZIP_STORED if path.suffix.lower() in {".gz", ".xlsx", ".zip"} else zipfile.ZIP_DEFLATED
            zf.write(path, rel, compress_type=compression, compresslevel=6 if compression == zipfile.ZIP_DEFLATED else None)
    with zipfile.ZipFile(out_zip, "r", allowZip64=True) as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"ZIP CRC failure: {bad}")


def copy_self(package: Path) -> None:
    scripts = package / "scripts"
    scripts.mkdir(exist_ok=True)
    shutil.copy2(Path(__file__), scripts / "build_lexicon.py")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--source-url", required=True)
    parser.add_argument("--source-headers", type=Path)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--target", type=int, default=TARGET)
    args = parser.parse_args()

    build_root = args.out.resolve()
    build_root.mkdir(parents=True, exist_ok=True)
    work = build_root / "work"
    package = build_root / "English_Lexicon_Kaikki_Wiktionary_Verified"
    for path in (work, package):
        if path.exists():
            shutil.rmtree(path)
        path.mkdir(parents=True)
    (package / "data" / "million").mkdir(parents=True)
    (package / "data" / "full").mkdir(parents=True)
    (package / "audit").mkdir(parents=True)
    (package / "proof").mkdir(parents=True)
    (package / "licenses").mkdir(parents=True)

    started = utc_now()
    source_hash = sha256_file(args.source)
    source_size = args.source.stat().st_size
    db_path = work / "lexicon.sqlite"
    db = init_db(db_path)
    process_counts = process_jsonl(args.source, db, build_root / "progress.json")

    cache = work / "recon"
    cache.mkdir(parents=True)
    recon_errors: list[str] = []
    try:
        dwyl = download_recon_dwyl(cache)
    except Exception as exc:
        dwyl = set()
        recon_errors.append(f"DWYL: {exc}")
    try:
        webster, webster_url = download_recon_webster(cache)
    except Exception as exc:
        webster = set()
        webster_url = "discovery_failed"
        recon_errors.append(f"Webster: {exc}")
    recon_counts = apply_recon(db, dwyl, webster)

    chosen = mark_million(db, args.target)
    full_count = db.execute("SELECT COUNT(*) FROM accepted").fetchone()[0]
    target_met = chosen == args.target

    full_export_count, full_letters = export_partitions(db, package / "data" / "full", "")
    million_export_count, million_letters = export_partitions(db, package / "data" / "million", "WHERE in_million=1")
    rejected_count = export_rejected(db, package / "audit" / "rejected_unique.csv.gz")

    reason_counts = dict(db.execute("SELECT reason,count FROM reason_counts ORDER BY count DESC"))
    full_pos = dict(db.execute(
        "SELECT pos,COUNT(*) FROM accepted GROUP BY pos ORDER BY COUNT(*) DESC"
    ))
    million_classes = dict(db.execute(
        "SELECT entry_class,COUNT(*) FROM accepted WHERE in_million=1 GROUP BY entry_class ORDER BY COUNT(*) DESC"
    ))

    metadata = {
        "build_version": VERSION,
        "started_utc": started,
        "completed_utc": utc_now(),
        "source_name": SOURCE_NAME,
        "source_url": args.source_url,
        "source_sha256": source_hash,
        "source_bytes": source_size,
        "source_lines": process_counts.get("source_lines", 0),
        "full_accepted_unique": full_count,
        "million_target": args.target,
        "million_rows": chosen,
        "target_met": target_met,
        "rejected_unique": rejected_count,
        "acceptance_rule": "English single-token lexical item with independent gloss or explicit form/alternative relation",
        "excluded": sorted(EXCLUDED_POS) + sorted(HARD_REJECT_TAG_FRAGMENTS),
        "recon_policy": "DWYL and Webster membership are comparison flags only; never validators",
        "recon": recon_counts,
        "recon_errors": recon_errors,
    }

    workbook_path = package / f"English_Lexicon_{chosen}_Verified.xlsx"
    write_workbook(db, workbook_path, metadata, chosen)

    source_headers_text = ""
    if args.source_headers and args.source_headers.exists():
        source_headers_text = args.source_headers.read_text(encoding="utf-8", errors="replace")
        shutil.copy2(args.source_headers, package / "proof" / "source_http_headers.txt")
    (package / "proof" / "source_sha256.txt").write_text(f"{source_hash}  {args.source.name}\n", encoding="utf-8")
    (package / "proof" / "manifest.json").write_text(json.dumps({
        **metadata,
        "process_counts": process_counts,
        "full_letter_counts": full_letters,
        "million_letter_counts": million_letters,
        "full_pos_counts": full_pos,
        "million_class_counts": million_classes,
        "rejection_reason_counts": reason_counts,
        "webster_recon_url": webster_url,
        "dwyl_recon_url": DWYL_URL,
        "source_http_headers": source_headers_text,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    (package / "README.md").write_text(f"""# Verified English Lexical Forms — Kaikki/Wiktionary\n\n## Result\n\n- Full accepted unique inventory: **{full_count:,}**\n- Deterministic target inventory: **{chosen:,}**\n- Requested one-million target met: **{target_met}**\n- Rejected unique candidates: **{rejected_count:,}**\n\n## Definition of verified\n\nEvery accepted row is a unique normalized spelling from the pinned Kaikki English source and has either an independent gloss or an explicit source-provided form-of/alternative-of relationship. Proper names, phrases, affixes, symbols, punctuation, documented misspellings, eye-dialect forms, reconstructions, and taxonomic names are excluded.\n\nThis package does **not** claim one million independent lemmas or every possible English word. It claims source-attested English lexical forms complete against the recorded source snapshot and rules.\n\nDWYL and Webster fields are reconnaissance only and never determine acceptance.\n\n## Contents\n\n- `English_Lexicon_{chosen}_Verified.xlsx` — review workbook\n- `data/million/` — deterministic target inventory, A–Z compressed CSV\n- `data/full/` — all accepted rows, A–Z compressed CSV\n- `audit/rejected_unique.csv.gz` — rejected candidates and reasons\n- `proof/manifest.json` — source, rules, counts, and run metadata\n- `proof/file_hashes.sha256` — package file hashes\n- `proof/verification_report.json` — independent consistency checks\n- `scripts/build_lexicon.py` — exact build code\n""", encoding="utf-8")

    (package / "licenses" / "NOTICE.md").write_text(
        "Kaikki data is extracted from Wiktionary and carries the applicable Wiktionary licensing requirements. "
        "Review the source site and Wiktionary licensing terms before redistribution. DWYL and Webster datasets are "
        "reconnaissance inputs only; their upstream licenses remain applicable.\n", encoding="utf-8")
    copy_self(package)

    unique_report = verify_unique(db, chosen)
    workbook_report = verify_workbook(workbook_path, chosen)
    csv_million_count = sum(count_gzip_csv_rows(p) for p in (package / "data" / "million").glob("*.csv.gz"))
    csv_full_count = sum(count_gzip_csv_rows(p) for p in (package / "data" / "full").glob("*.csv.gz"))
    if csv_million_count != chosen or csv_full_count != full_count:
        raise RuntimeError({"csv_million": csv_million_count, "chosen": chosen,
                            "csv_full": csv_full_count, "full": full_count})
    verification = {
        "verified_at_utc": utc_now(),
        "status": "PASS" if target_met else "PASS_SOURCE_BUILD_TARGET_NOT_MET",
        "target_met": target_met,
        "database_checks": unique_report,
        "workbook_checks": workbook_report,
        "csv_million_rows": csv_million_count,
        "csv_full_rows": csv_full_count,
        "xlsx_sha256": sha256_file(workbook_path),
        "source_sha256": source_hash,
        "recon_not_used_for_acceptance": True,
    }
    (package / "proof" / "verification_report.json").write_text(
        json.dumps(verification, ensure_ascii=False, indent=2), encoding="utf-8")
    hashes = write_file_hashes(package)
    verification["hashed_package_files"] = len(hashes)
    (package / "proof" / "verification_report.json").write_text(
        json.dumps(verification, ensure_ascii=False, indent=2), encoding="utf-8")
    write_file_hashes(package)

    out_zip = build_root / f"English_Lexicon_Kaikki_Wiktionary_Verified_{chosen}.zip"
    make_zip(package, out_zip)
    zip_hash = sha256_file(out_zip)
    (build_root / (out_zip.name + ".sha256.txt")).write_text(f"{zip_hash}  {out_zip.name}\n", encoding="utf-8")
    summary = {
        **metadata,
        "package_zip": out_zip.name,
        "package_zip_sha256": zip_hash,
        "package_zip_bytes": out_zip.stat().st_size,
        "workbook": workbook_path.name,
        "workbook_sha256": sha256_file(workbook_path),
        "verification": verification,
    }
    (build_root / "BUILD_SUMMARY.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    db.close()
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
